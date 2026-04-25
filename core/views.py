from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from .models import Sale, Store, Promotion, Competition, Announcement, Product, Category
from django.views.generic import ListView, TemplateView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic.edit import CreateView
from django.urls import reverse_lazy
from .forms import SaleForm
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.contrib.auth import get_user_model
from django import forms

User = get_user_model()

class StoreAccessMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.role in ['admin', 'promoter']

class CompetitionListView(LoginRequiredMixin, ListView):
    model = Competition
    template_name = 'competition/competition_list.html'
    context_object_name = 'entries'
    paginate_by = 10  # Προσθήκη Pagination

    def get_queryset(self):
        queryset = Competition.objects.all().select_related('store', 'salesperson')
        
        if self.request.user.role == 'promoter':
            queryset = queryset.filter(salesperson=self.request.user)

        # Φίλτρο αναζήτησης
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(store__name__icontains=q) | queryset.filter(comments__icontains=q)
            
        return queryset.order_by('-date')

class CompetitionCreateView(LoginRequiredMixin, CreateView):
    model = Competition
    fields = ['store', 'comments']
    template_name = 'competition/competition_form.html'
    success_url = reverse_lazy('competition-list')

    def form_valid(self, form):
        form.instance.salesperson = self.request.user
        return super().form_valid(form)
    
def export_competition_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Competition_Report.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Ανταγωνισμός"

    headers = ['Ημερομηνία', 'Κατάστημα', 'Σχόλια / Παρατηρήσεις', 'Promoter']
    ws.append(headers)

    # Styling
    for cell in ws[1]:
        cell.font = Font(bold=True)

    entries = Competition.objects.all().select_related('store', 'salesperson')
    if request.user.role == 'promoter':
        entries = entries.filter(salesperson=request.user)

    for entry in entries:
        ws.append([
            entry.date.strftime("%d/%m/%Y"),
            entry.store.name,
            entry.comments,
            entry.salesperson.get_full_name() or entry.salesperson.username
        ])

    wb.save(response)
    return response

@login_required
def dashboard_view(request):
    user = request.user
    
    # Αρχικά Querysets
    sales_qs = Sale.objects.all()
    promos_qs = Promotion.objects.all()

    # Αν είναι Promoter, βλέπει ΜΟΝΟ τα δικά του (spec requirement!)
    if user.role == 'promoter':
        sales_qs = sales_qs.filter(salesperson=user)
        promos_qs = promos_qs.filter(salesperson=user)

    # Υπολογισμός Συνόλων (Sum)
    total_sales = sales_qs.aggregate(Sum('quantity'))['quantity__sum'] or 0
    total_promos = promos_qs.aggregate(Sum('quantity'))['quantity__sum'] or 0

    # Τελευταίες 10 πωλήσεις
    recent_sales = sales_qs.order_by('-date')[:10]

    context = {
        'total_sales': total_sales,
        'total_promos': total_promos,
        'recent_sales': recent_sales,
    }
    
    return render(request, 'dashboard.html', context)

class SaleListView(LoginRequiredMixin, ListView):
    model = Sale
    template_name = 'sales/sale_list.html'
    context_object_name = 'sales'
    paginate_by = 50

    def get_queryset(self):
        queryset = Sale.objects.all().select_related('product', 'store', 'salesperson')
        
        # 1. Φιλτράρισμα βάσει ρόλου (το κρατάμε!)
        if self.request.user.role == 'promoter':
            queryset = queryset.filter(salesperson=self.request.user)

        # 2. Αναζήτηση (Search)
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(store__name__icontains=q) | queryset.filter(product__name__icontains=q)

        # 3. Φιλτράρισμα Ημερομηνίας
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        if start_date and end_date:
            queryset = queryset.filter(date__range=[start_date, end_date])

        return queryset.order_by('-date', '-created_at')
    
class SaleCreateView(LoginRequiredMixin, CreateView):
    model = Sale
    form_class = SaleForm
    template_name = 'sales/sale_form.html'
    success_url = reverse_lazy('sale-list')

    def form_valid(self, form):
        # Αυτόματη ανάθεση του συνδεδεμένου χρήστη ως salesperson
        form.instance.salesperson = self.request.user
        return super().form_valid(form)

class SalesStatusView(UserPassesTestMixin, TemplateView):
    template_name = 'dashboard/sales_status.html'

    def test_func(self):
        # Επιτρέπει πρόσβαση μόνο σε admin και promoter (όχι σε clients)
        return self.request.user.role in ['admin', 'promoter']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        sales = Sale.objects.all().select_related('product', 'store', 'salesperson')

        # 1. Εφαρμογή Φίλτρων
        store_id = self.request.GET.get('store')
        promoter_id = self.request.GET.get('promoter')
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')

        if store_id:
            sales = sales.filter(store_id=store_id)
        if promoter_id:
            sales = sales.filter(salesperson_id=promoter_id)
        if start_date:
            sales = sales.filter(date__gte=start_date)
        if end_date:
            sales = sales.filter(date__lte=end_date)

        # 2. Aggregation ανά Προϊόν
        product_performance = (
            sales.values('product__name')
            .annotate(total_qty=Sum('quantity'))
            .order_by('-total_qty')
        )

        context.update({
            'product_performance': product_performance,
            'stores': Store.objects.all(),
            'promoters': User.objects.filter(role='promoter'),
        })
        return context

def export_sales_excel(request):
    # Δημιουργία του response με το σωστό header για Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Sales_Report.xlsx"'

    # Δημιουργία αρχείου
    wb = Workbook()
    ws = wb.active
    ws.title = "Πωλήσεις"

    # Ορισμός κεφαλίδων
    headers = ['Ημερομηνία', 'Κατάστημα', 'Προϊόν', 'Ποσότητα', 'Promoter']
    ws.append(headers)

    # Μορφοποίηση κεφαλίδων (Bold & Center)
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Φιλτράρισμα δεδομένων (Role-based)
    sales = Sale.objects.all().select_related('store', 'product', 'salesperson').order_by('-date')
    
    if request.user.role == 'promoter':
        sales = sales.filter(salesperson=request.user)

    # Προσθήκη δεδομένων
    for sale in sales:
        ws.append([
            sale.date.strftime("%d/%m/%Y"),
            sale.store.name,
            sale.product.name,
            sale.quantity,
            sale.salesperson.get_full_name() or sale.salesperson.username
        ])

    # Αυτόματη ρύθμιση πλάτους στηλών
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2

    wb.save(response)
    return response

class ProductListView(UserPassesTestMixin, ListView):
    model = Product
    template_name = 'products/product_list.html'
    context_object_name = 'products'

    def test_func(self):
        return self.request.user.role in ['admin', 'promoter']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = Category.objects.all()
        return context

# View για τη δημιουργία κατηγορίας (θα καλείται από το Modal)
def add_category(request):
    if request.method == 'POST' and request.user.role != 'client':
        name = request.POST.get('name')
        if name:
            Category.objects.create(name=name)
    return redirect('product-list')

class StoreListView(LoginRequiredMixin, StoreAccessMixin, ListView):
    model = Store
    template_name = 'stores/store_list.html'
    context_object_name = 'stores'

    def get_queryset(self):
        # Δείξε μόνο αυτά που δεν έχουν διαγραφεί (soft delete)
        return Store.objects.filter(is_active=True)

class StoreCreateView(LoginRequiredMixin, StoreAccessMixin, CreateView):
    model = Store
    fields = ['name', 'address', 'city'] # Προσάρμοσε τα πεδία ανάλογα με το μοντέλο σου
    template_name = 'stores/store_form.html'
    success_url = reverse_lazy('store-list')

class StoreUpdateView(LoginRequiredMixin, StoreAccessMixin, UpdateView):
    model = Store
    fields = ['name', 'address', 'city']
    template_name = 'stores/store_form.html'
    success_url = reverse_lazy('store-list')

class StoreDeleteView(LoginRequiredMixin, StoreAccessMixin, DeleteView):
    model = Store
    template_name = 'stores/store_confirm_delete.html'
    success_url = reverse_lazy('store-list')


class AdminOnlyMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.role == 'admin'

class UserListView(LoginRequiredMixin, AdminOnlyMixin, ListView):
    model = User
    template_name = 'users/user_list.html'
    context_object_name = 'users_list'

    def get_queryset(self):
        # Δείχνουμε μόνο τους ενεργούς χρήστες (soft delete)
        return User.objects.filter(is_active=True).order_by('-date_joined')

class UserCreateView(LoginRequiredMixin, AdminOnlyMixin, CreateView):
    model = User
    template_name = 'users/user_form.html'
    fields = ['username', 'email', 'first_name', 'last_name', 'role', 'password']
    success_url = reverse_lazy('user-list')

    def form_valid(self, form):
        # Χρειάζεται προσοχή για να γίνει σωστά hash το password
        user = form.save(commit=False)
        user.set_password(form.cleaned_data['password'])
        user.save()
        return super().form_valid(form)
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Κάνουμε το password field να φαίνεται ως password input
        form.fields['password'].widget = forms.PasswordInput(attrs={'class': 'rounded-xl border-gray-200 w-full'})
        return form

class UserUpdateView(LoginRequiredMixin, AdminOnlyMixin, UpdateView):
    model = User
    fields = ['first_name', 'last_name', 'role', 'email']
    template_name = 'users/user_form.html'
    success_url = reverse_lazy('user-list')

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Προσθέτουμε Tailwind classes στα πεδία
        for field in form.fields.values():
            field.widget.attrs.update({'class': 'rounded-xl border-gray-200 w-full focus:ring-indigo-500 focus:border-indigo-500'})
        return form

class UserDeleteView(LoginRequiredMixin, AdminOnlyMixin, DeleteView):
    model = User
    template_name = 'users/user_confirm_delete.html'
    success_url = reverse_lazy('user-list')

    def delete(self, request, *args, **kwargs):
        user = self.get_object()
        user.is_active = False # Soft delete
        user.save()
        return redirect(self.success_url)